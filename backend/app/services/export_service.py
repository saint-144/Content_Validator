"""
Excel Export Service
Generates .xlsx reports for validation results
"""

import io
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


VERDICT_COLORS = {
    "appropriate": "10B981",
    "escalate": "EF4444",
    "need_review": "F59E0B"
}
HEADER_COLOR = "1E3A5F"


def _apply_header(ws, row: int, cols: list):
    for col_idx, header in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            bottom=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF")
        )


def _verdict_fill(verdict: str) -> PatternFill:
    color = VERDICT_COLORS.get(verdict, "9CA3AF")
    return PatternFill("solid", fgColor=color)


def export_validations_to_excel(validations_data: List[dict]) -> bytes:
    """
    Export validation results to Excel.
    Returns bytes of the .xlsx file.
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Validation Summary"
    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 25
    ws_summary.column_dimensions["C"].width = 35
    ws_summary.column_dimensions["D"].width = 20
    ws_summary.column_dimensions["E"].width = 20
    ws_summary.column_dimensions["F"].width = 35
    ws_summary.column_dimensions["G"].width = 18
    ws_summary.column_dimensions["H"].width = 15
    ws_summary.column_dimensions["I"].width = 15

    # Title
    ws_summary.merge_cells("A1:I1")
    title_cell = ws_summary["A1"]
    title_cell.value = f"Content Validation Report — Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    title_cell.font = Font(bold=True, size=13, color="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center")
    ws_summary.row_dimensions[1].height = 30

    headers = [
        "Report Ref", "Post Timestamp", "Post Description / Caption",
        "Template Used", "Suspected Match File(s)", "Exact Pixel Match?",
        "MCC Compliant?", "Action", "Validated At"
    ]
    _apply_header(ws_summary, 2, headers)
    ws_summary.row_dimensions[2].height = 25

    for row_idx, v in enumerate(validations_data, 3):
        # Suspected match files
        suspected = [
            m["template_file_name"]
            for m in v.get("matches", [])
            if m.get("is_suspected_match")
        ]
        suspected_str = ", ".join(suspected) if suspected else "No Match"
        match_display = f"{v.get('template_name', '')} > {suspected_str}" if suspected else "No Match Found"

        exact = any(m.get("is_exact_pixel_match") for m in v.get("matches", []))
        mcc = v.get("mcc_compliant")
        verdict = v.get("overall_verdict", "need_review")
        action_map = {"appropriate": "Appropriate ✓", "escalate": "Escalate 🚨", "need_review": "Need to Review ⚠️"}

        row_data = [
            v.get("report_ref", ""),
            v.get("post_timestamp", "") or "",
            (v.get("post_description", "") or "")[:200],
            v.get("template_name", ""),
            match_display,
            "Yes" if exact else "No",
            "Yes" if mcc else ("No" if mcc is False else "N/A"),
            action_map.get(verdict, verdict),
            str(v.get("created_at", ""))[:19]
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(
                bottom=Side(style="thin", color="E5E7EB"),
                right=Side(style="thin", color="E5E7EB")
            )

        # Color-code the Action column
        action_cell = ws_summary.cell(row=row_idx, column=8)
        action_cell.fill = _verdict_fill(verdict)
        action_cell.font = Font(color="FFFFFF", bold=True, size=9)
        action_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws_summary.row_dimensions[row_idx].height = 40

    # ── Sheet 2: Detailed Matches ─────────────────────────────────────────────
    ws_detail = wb.create_sheet("Detailed Match Results")
    ws_detail.column_dimensions["A"].width = 15
    ws_detail.column_dimensions["B"].width = 20
    ws_detail.column_dimensions["C"].width = 35
    ws_detail.column_dimensions["D"].width = 18
    ws_detail.column_dimensions["E"].width = 18
    ws_detail.column_dimensions["F"].width = 18
    ws_detail.column_dimensions["G"].width = 15
    ws_detail.column_dimensions["H"].width = 15
    ws_detail.column_dimensions["I"].width = 50

    detail_headers = [
        "Report Ref", "Template Name", "Template File", "LLM Score (%)",
        "Pixel Score (%)", "Overall Score (%)", "Suspected Match?",
        "Exact Pixel Match?", "Match Reasoning"
    ]
    _apply_header(ws_detail, 1, detail_headers)

    detail_row = 2
    for v in validations_data:
        for m in v.get("matches", []):
            row_data = [
                v.get("report_ref", ""),
                v.get("template_name", ""),
                m.get("template_file_name", ""),
                float(m.get("llm_similarity_score", 0)),
                float(m.get("pixel_similarity_score", 0)),
                float(m.get("overall_similarity_score", 0)),
                "Yes" if m.get("is_suspected_match") else "No",
                "Yes" if m.get("is_exact_pixel_match") else "No",
                m.get("match_reasoning", "")[:300]
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_detail.cell(row=detail_row, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))

            if m.get("is_suspected_match"):
                ws_detail.cell(row=detail_row, column=7).fill = PatternFill("solid", fgColor="FEF3C7")
            if m.get("is_exact_pixel_match"):
                ws_detail.cell(row=detail_row, column=8).fill = PatternFill("solid", fgColor="D1FAE5")

            ws_detail.row_dimensions[detail_row].height = 35
            detail_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
